# Codé par THEVAKUMARAN Krishnarash, LASNAMI Sara
# Optimisation Combinatoire
# Bin Packing (Methode First Fit Decreasing)

import openpyxl
import xlsxwriter
import numpy as np
import random

""" On implémente la classe Sac """
class Sac(object):
    def __init__(self):
        # La liste qui va contenir les objets qui vont rentrer dans le sac
        self.objets = []
        self.poids_total = 0

    # La fonction ajout_objets permet d'ajouter un objet dans le sac
    def ajout_objets(self, objet):
        self.objets.append(objet)
        self.poids_total += objet

""" Algorithme de la méthode First Fit """
def first_fit(liste_objets, capacite):
    sacs = []
    liste_aleatoire = np.random.permutation(liste_objets)  # list containing initial items in a random order
    liste_objets = liste_aleatoire.tolist()

    for objet_liste_random in liste_objets:
        # pour chaque objets, nous cherchons s'il y a un sac libre où l'objet peut être placé.
        for objets in sacs:
            # Si le poids_total de notre sac et le poids de l'objet ne depasse pas la capacité de notre sac
            if objets.poids_total + objet_liste_random <= capacite:
                # nous ajoutons l'objet dans le sac avec la fonction ajout_objets
                objets.ajout_objets(objet_liste_random)
                break
        # Sinon : c'est à dire s'il n'y a pas de sac libre où l'objet peut être placé.
        else:
            # Nous ouvrons donc un nouveau sac et y ajoutons l'objet.
            objets = Sac()
            objets.ajout_objets(objet_liste_random)
            sacs.append(objets)

    return sacs

""" Algorithme de la méthode First Fit Decreasing """
def first_fit_decreasing(liste_objets, capacite_sac):
    # Trie la liste des objets dans un ordre décroissant avec la methode sorted et reverse = True.
    liste_decroissant = sorted(liste_objets, reverse=True)
    sacs =[]
    for objet_liste_dec in liste_decroissant:
        # pour chaque objets, nous cherchons s'il y a un sac libre où l'objet peut être placé.
        for objets in sacs:
            # Si le poids_total de notre sac et le poids de l'objet ne depasse pas la capacité de notre sac
            if objets.poids_total + objet_liste_dec <= capacite_sac:
                # nous ajoutons l'objet dans le sac avec la fonction ajout_objets
                objets.ajout_objets(objet_liste_dec)
                break
        # Sinon : c'est à dire s'il n'y a pas de sac libre où l'objet peut être placé.
        else:
            # Nous ouvrons donc un nouveau sac et y ajoutons l'objet.
            objets = Sac()
            objets.ajout_objets(objet_liste_dec)
            sacs.append(objets)
    return sacs

""" Cette fonction prend un fichier (excel de preference) qu'on renseigne en parametre
et renvoie la liste des objets et la capacité du sac. """
def extraction_fichier(fichier):
    fichier1 = openpyxl.load_workbook(fichier)
    feuille = fichier1.active
    liste_objets = []
    # Recupere le nombre d'objet qui est dans la 2eme ligne de la 1ere colonne
    nbre_objets = feuille.cell(row = 2, column = 1).value
    # Recupere la capacite du sac qui est dans la 2eme ligne de la 2eme colonne
    capacite = feuille.cell(row = 2, column = 2).value
    for i in range(2,nbre_objets+2):
        # Recupere le poids des objets de la 3eme colonne
        case_objet = feuille.cell(row = i, column = 3)
        liste_objets.append(case_objet.value)
    return liste_objets,capacite

""" Cette fonction prend la liste des objets et la capacité du sac et renvoi la solution optimal des majorants """
def solution_optimal(liste_objets, capacite_sac):
    poids_total = 0
    for objet_liste in liste_objets:
        # Nous ouvrons donc un nouveau sac et y ajoutons l'objet et chaque ajout d'objet, le poids de l'objet s'ajoutera au poids total
        objets = Sac()
        objets.ajout_objets(objet_liste)
        poids_total += objets.poids_total
    # Solution Optimal est le poids total divise par la capacite du sac
    solution = poids_total/(capacite_sac)
    return solution

""" Cette fonction genere un fichier excel dans notre cas, pour inserer la capacite, le nombre d'objet et la liste des objets aleatoire
 pour ensuite savoir les nombres de sac utilise avec la methode ffd"""
def generation_fichier(fichier):
    # Creation d'un fichier excel et d'une feuille excel
    workbook = xlsxwriter.Workbook(fichier)
    worksheet = workbook.add_worksheet("Resultat")

    # Instantiation des valeurs (ligne, capacite, nombre d'objets, liste d'objets, solution optimal)
    ligne = 1
    capacite = random.randint(1,200)
    nbre_objet = random.randint(1,100)
    liste_objet = [random.randint(0,capacite) for p in range(0,nbre_objet)]
    bin_package_first_fit_decreasing = first_fit_decreasing(liste_objet, capacite)
    bin_package_first_fit_random = first_fit(liste_objet, capacite)
    solution = solution_optimal(liste_objet, capacite)

    # Ecriture des valeurs dans la feuille de calcul d'excel
    worksheet.write("A1", "Nombres d'objets")
    worksheet.write("A2", nbre_objet)
    worksheet.write("B1", "Capacité du sac")
    worksheet.write("B2", capacite)
    worksheet.write("C1", "Poids")

    # Ecriture des listes d'objets en colonne sur l'excel
    for item in liste_objet:
        worksheet.write(ligne, 2, item)
        ligne += 1

    # Ecriture des resultats numeriques sur l'excel
    worksheet.write("E1", "Solution optimal")
    worksheet.write("E2", round(solution))
    worksheet.write("F1", "Sac FFD")
    worksheet.write("F2", len(bin_package_first_fit_decreasing))
    worksheet.write("G1", "Efficacite de l'algo FFD")
    worksheet.write("G2", round((len(bin_package_first_fit_decreasing) - solution)))
    worksheet.write("H1", "Sac FF")
    worksheet.write("H2", len(bin_package_first_fit_random))
    worksheet.write("I1", "Efficacite de l'algo FF")
    worksheet.write("I2", round((len(bin_package_first_fit_random) - solution)))

    # Fermeture de l'excel
    workbook.close()

#Appel de fonction de generation de fichier excel (Resultats numeriques sur ces excels)
generation_fichier("DonneeTest_FFD1.xlsx")
generation_fichier("DonneeTest_FFD2.xlsx")
generation_fichier("DonneeTest_FFD3.xlsx")
generation_fichier("DonneeTest_FFD4.xlsx")
generation_fichier("DonneeTest_FFD5.xlsx")
generation_fichier("DonneeTest_FFD6.xlsx")
generation_fichier("DonneeTest_FFD7.xlsx")
generation_fichier("DonneeTest_FFD8.xlsx")
generation_fichier("DonneeTest_FFD9.xlsx")
generation_fichier("DonneeTest_FFD10.xlsx")

########################################################################################################################
###################### Affichage sur console des resultats numeriques du premiers excels (Exemple) ###############################

print("\nExemple affichage console de l'algorithme FFD sur le premier excel")
resultat = extraction_fichier("DonneeTest_FFD1.xlsx")
liste_objet, capacite = resultat[0], resultat[1]
bin_package_first_fit_decreasing = first_fit_decreasing(liste_objet, capacite)
bin_package_first_fit_random = first_fit(liste_objet, capacite)
solution = solution_optimal(liste_objet, capacite)

print("La solution optimale : ", round(solution), "sacs")

print("Le nombre de sac utilise avec la méthode First Fit Decreasing : ", len(bin_package_first_fit_decreasing), "sacs")
print("L'efficacite de l'algorithme approche de la methode bin packing ffd :", round((len(bin_package_first_fit_decreasing) - solution)))

print("\nLe nombre de sac utilise avec la méthode First Fit : ", len(bin_package_first_fit_random), "sacs")
print("L'efficacite de l'algorithme approche de la methode bin packing ff :", round((len(bin_package_first_fit_random) - solution)))




